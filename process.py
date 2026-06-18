import json, os, glob
from openpyxl import load_workbook

def get_val(row, headers, *names):
    for n in names:
        if n in headers:
            v = row[headers[n]]
            return '' if v is None else str(v).replace('.0','').strip()
    return ''

def get_num(row, headers, *names):
    for n in names:
        if n in headers:
            v = row[headers[n]]
            try: return int(float(v)) if v else 0
            except: return 0
    return 0

def get_date(row, headers, *names):
    from datetime import datetime
    for n in names:
        if n in headers:
            v = row[headers[n]]
            if not v: continue
            if isinstance(v, datetime): return v.strftime('%d.%m.%Y')
            s = str(v).strip()
            if '.' in s: return s
            if '-' in s: parts=s[:10].split('-'); return f"{parts[2]}.{parts[1]}.{parts[0]}" if len(parts)==3 else s
    return ''

def read_xlsx(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    headers = {str(h).strip(): i for i, h in enumerate(rows[0]) if h}
    return [rows, headers]

# SAP dosyasını bul
sap_files = glob.glob('uploads/sap_*.xlsx') + glob.glob('uploads/SAP_*.xlsx') + glob.glob('uploads/sap_*.xls')
omnia_files = glob.glob('uploads/omnia_*.xlsx') + glob.glob('uploads/OMNIA_*.xlsx') + glob.glob('uploads/omnia_*.xls')

orders = []

if sap_files:
    result = read_xlsx(sorted(sap_files)[-1])
    if result:
        rows, headers = result
        omnia_map = {}
        if omnia_files:
            r2 = read_xlsx(sorted(omnia_files)[-1])
            if r2:
                orows, oheaders = r2
                for row in orows[1:]:
                    key = get_val(row, oheaders, 'Sistem No', 'SistemNo')
                    if key: omnia_map[key] = (row, oheaders)

        for row in rows[1:]:
            ship_to = get_val(row, headers, 'Ship To Name')
            if 'YENİ MA' not in ship_to.upper(): continue
            del_doc = get_val(row, headers, 'Delivery Document Number')
            omn_row, omn_h = omnia_map.get(del_doc, (None, None))
            orders.append({
                'p': get_val(row, headers, 'Customer PO Number'),
                'd': del_doc,
                'dp': ship_to,
                'sh': get_val(row, headers, 'Ship to City').upper().strip(),
                'u': get_val(row, headers, 'Material Description'),
                'sk': get_num(row, headers, 'Order Quantity in Sales Unit'),
                'tk': get_num(row, headers, 'Picked Quantity', 'Delivery Quantity'),
                'bt': get_date(row, headers, 'Billing date'),
                'gt': get_date(row, headers, 'Actual Goods Movement Date'),
                'oc': get_date(row, headers, 'Sales Order Creation date', 'Order Creation Date'),
                'pd': get_date(row, headers, 'Proof of Delivery Date', 'ProofOfDeliveryDate'),
                'a': get_val(omn_row, omn_h, 'ATF Durum') if omn_row else '',
                'pl': get_val(omn_row, omn_h, 'Plaka') if omn_row else '',
                'ot': get_date(omn_row, omn_h, 'Teslim Tarihi') if omn_row else '',
            })

from datetime import datetime
payload = {'orders': orders, 'updatedAt': datetime.now().strftime('%d.%m.%Y %H:%M')}
with open('data.json', 'w', encoding='utf-8') as f:
    json.dump(payload, f, ensure_ascii=False)
print(f"✅ {len(orders)} sipariş yazıldı.")
